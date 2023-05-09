import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import openpyxl
import cloudscraper as cs

from datetime import timedelta
from bs4 import BeautifulSoup as bs
from sklearn.metrics import roc_curve, auc
from Probability import probability

k = np.arange(5.5, 6.5, 0.01)
hta = np.round(np.arange(-500., 500., 10), 2)
mvma = np.round(np.arange(0.1, 1.1, 0.01), 2)
mvmb = np.round(np.arange(0.5, 1.1, 0.01), 2)
acf = np.arange(2.02, 2.09, 0.01)
acm = np.round(np.arange(0.0007, 0.0013, 0.0001), 4)

URL = 'https://www.hockey-reference.com/leagues/NHL_2023_games.html'

args = {'browser': 'firefox', 'desktop': True, 'platform': 'windows'}

scraper = cs.create_scraper(browser=args)
response = scraper.get(URL)

soup = bs(response.text, 'html.parser')
tex = soup.find('table').find_all('tr')
wb_name = 'Main.xlsx'

wb = openpyxl.Workbook()
wb.save(wb_name)
wb.close()

yest_games = []
for i in range(200):
    start = pd.to_datetime('2022-10-06', format='%Y-%m-%d') + timedelta(days=i)
    ##################
    today_date = start.strftime('%Y-%m-%d')
    for _, t in enumerate(tex):
        if str(t.find('th').text) == str(today_date):
            yest_games.append((t.find_all('td')[0].text, t.find_all('td')[1].text, t.find_all('td')[2].
                               text, t.find_all('td')[3].text))

a = 6.45
b = 0
c = 0.89
d = 0.9
e = 2.05
f = 0.001

data = pd.DataFrame(columns=['Team', 'Elo'])
wb = openpyxl.load_workbook('Elo_test.xlsx')
ws = wb.active

for j in range(2, 34):
    data = data.append({'Team': ws[f'A{j}'].value, 'Elo': ws[f'B{j}'].value}, ignore_index=True)

elo = data.set_index('Team', drop=True).to_dict()['Elo']

wb.close()

auc_scores = []
games_num = []
for i in range(3,31):
    games = yest_games[:int(len(yest_games)*i/30)]
    games_num.append(len(games))
    probability(K=a, HTA=b, MVM_A=c, MVM_B=d, ACF=e, ACM=f, wb_name=wb_name, sheet_num=i, yest_games=games,
                elo_sub=elo)

    data = pd.read_excel('Main.xlsx', sheet_name=f'sheet{i}', header=[2])

    fpr, tpr, thresholds = roc_curve(data['home_win'], data['home_prob'])
    roc_auc = auc(fpr, tpr)
    auc_scores.append(roc_auc)

#################################

# nhl = pd.read_csv('nhl_elo.csv')
# col = nhl.columns
#
# nhl['date'] = pd.to_datetime(nhl['date'])
# nhl = nhl.query('season == 2023')
# col_to_keep = ['date', 'neutral', 'home_team','away_team','home_team_pregame_rating', 'away_team_pregame_rating',
#                'home_team_winprob', 'away_team_winprob', 'overtime_prob','home_team_score', 'away_team_score',
#                'home_team_postgame_rating', 'away_team_postgame_rating']
#
# nhl = nhl[col_to_keep].reset_index(drop=True).dropna()
#
# t = 0
# p = 0
# y_true = []
# y_score = []
# for i, r in nhl.iterrows():
#     t = int((r['home_team_score'] - r['away_team_score']) > 0)
#     p = r['home_team_winprob']
#     y_true.append(t)
#     y_score.append(p)
#
# games_num2 = []
# auc_scores2 = []
# for i in range(3, 31):
#     fpr, tpr, thresholds = roc_curve(y_true[:int(len(y_true)*i/30)], y_score[:int(len(y_true)*i/30)])
#     roc_auc = auc(fpr, tpr)
#     games_num2.append(int(len(y_true)*i/30))
#     auc_scores2.append(roc_auc)

##############

a = 6
b = 50
c = 0.6686
d = 0.8048
wb_name = 'Main1.xlsx'
wb = openpyxl.Workbook()
wb.save(wb_name)
wb.close()

auc_scores3 = []
for i in range(3, 31):
    games = yest_games[:int(len(yest_games)*i/30)]
    probability(K=a, HTA=b, MVM_A=c, MVM_B=d, ACF=e, ACM=f, wb_name=wb_name, sheet_num=i, yest_games=games,
                elo_sub=elo)

    data = pd.read_excel('Main1.xlsx', sheet_name=f'sheet{i}', header=[2])

    fpr, tpr, thresholds = roc_curve(data['home_win'], data['home_prob'])
    roc_auc = auc(fpr, tpr)
    auc_scores3.append(roc_auc)

######################

plt.style.use('seaborn-v0_8-deep')
plt.plot(games_num, auc_scores)
# plt.plot(games_num2, auc_scores2, color='orange')
plt.plot(games_num, auc_scores3, color='red')
plt.title("AUC Score Comparison")
plt.xlabel("Number of Games")
plt.ylabel("AUC Scores")
plt.scatter(games_num, auc_scores, color='darkblue', label='Optimization')
# plt.scatter(games_num2, auc_scores2, color='darkorange',label='FiveThirtyEight')
plt.scatter(games_num, auc_scores3, color='darkred',label='FTE_Advertised')
# plt.plot(fpr, tpr, color='darkorange', lw=2, label='ROC curve (area = %0.2f)' % roc_auc)
# plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
# plt.xlim([0.0, 1.0])
# plt.ylim([0.0, 1.05])
# plt.xlabel('False Positive Rate')
# plt.ylabel('True Positive Rate')
# plt.title('Receiver Operating Characteristic (ROC) Curve')
# plt.legend(loc="lower right")
# plt.savefig('ROC_Final')
plt.legend()
plt.tight_layout()
plt.savefig("Final AUC Scores2")
plt.show()



def my_func():
    sheet_num = 1
    scores = []

    data = pd.DataFrame(columns=['Team', 'Elo'])
    wb = openpyxl.load_workbook('Elo_test.xlsx')
    ws = wb.active

    for j in range(2, 34):
        data = data.append({'Team': ws[f'A{j}'].value, 'Elo': ws[f'B{j}'].value}, ignore_index=True)

    elo = data.set_index('Team', drop=True).to_dict()['Elo']

    wb.close()

    for f in acm:
        probability(K=a, HTA=b, MVM_A=c, MVM_B=d, ACF=e, ACM=f, wb_name=wb_name, sheet_num=sheet_num,
                    yest_games=yest_games, elo_sub=elo)

        data = pd.read_excel('Main.xlsx', sheet_name=f'sheet{sheet_num}', header=[2])

        sheet_num += 1

        fpr, tpr, thresholds = roc_curve(data['home_win'], data['home_prob'])
        roc_auc = auc(fpr, tpr)
        scores.append(roc_auc)
        print(f'{f} Done!')

    print(max(scores), scores.index(max(scores)) + 1)

    plt.plot(acm, scores, marker = 'o', color = 'red')
    plt.xlabel('ACM')
    plt.ylabel('AUC score')
    plt.title('ACM  -  AUC Scores')
    plt.savefig('ACM_AUC')
    plt.show()
    return


